"""
Find Connections (Basic) skill

Automates research of LinkedIn connections between target companies and key
contacts, populating the Excel "Company shortlist" spreadsheet.

Usage:
  python linkedin_excel_contact_skill.py "John Doe" https://linkedin.com/in/johndoe "Shortlisted companies.xlsx"
  python linkedin_excel_contact_skill.py "Acme Corp" https://linkedin.com/company/acme "Shortlisted companies.xlsx" --refresh
"""

import copy
import sys
import difflib
import os
from datetime import datetime
from typing import Callable, Dict, List, Optional, Tuple

from openpyxl.utils import get_column_letter

from chrome_bridge import CachedFetcher, plan_fetch_urls
from excel_utils import (
    SHEET_SCHEMA,
    ExcelManager,
    SheetStructure,
    create_sheet_from_schema,
    detect_sheet_structure,
    refresh_autofilter,
)
from linkedin_utils import LinkedInResearcher, find_connections_in_pages
from url_parser import LinkedInURLParser
from user_config import ConfigNotFoundError, UserConfig
from workspace import default_cache_path


class LinkedInExcelContactSkill:
    """Orchestrates LinkedIn research + Excel updates."""

    # Both defaults live on the schema, which is what the sheet writer uses too;
    # duplicating the literals here would let the two drift apart silently.
    DEFAULT_STATUS = SHEET_SCHEMA.default_status
    # New companies are rated P2, never P1: nothing should read as a top target
    # before a human has reviewed it. The rating itself stays a manual, analytical
    # act — the skill only ensures the column is never left blank.
    DEFAULT_PRIORITY = SHEET_SCHEMA.default_priority
    FUZZY_MATCH_THRESHOLD = 0.80  # difflib ratio for near-duplicate detection

    def __init__(
        self,
        config: UserConfig,
        fetch_fn: Optional[Callable[[str], str]] = None,
    ):
        self.config = config
        self.excel_manager = ExcelManager(config.excel_dir)
        self.researcher = LinkedInResearcher(fetch_fn=fetch_fn)
        self.url_parser = LinkedInURLParser()

        # Set by --apply so cleanup-on-success can delete the response cache.
        self.cache_path: Optional[str] = None
        # True/False after a save attempt; None before any save. Lets the caller
        # gate cache deletion on a confirmed successful write.
        self.last_save_ok: Optional[bool] = None

        # Role criteria come entirely from the per-user config — no built-ins.
        self.researcher.set_role_patterns(config.role_patterns())

    # ------------------------------------------------------------------ #
    # Public entry point                                                   #
    # ------------------------------------------------------------------ #

    def _open_sheet(self, linkedin_url: str, filename: str):
        """
        Validate inputs, load the workbook, and detect the sheet structure.

        Returns ``(profile_type, full_path, ws, structure)`` on success, or
        ``(None, error_message, None, None)`` on failure so callers can early-out.
        """
        if not filename or not filename.endswith(".xlsx"):
            return None, "Error: filename must be an .xlsx file", None, None

        profile_type, _ = self.url_parser.parse(linkedin_url)
        if not profile_type:
            return None, f"Error: could not parse LinkedIn URL: {linkedin_url}", None, None

        full_path = os.path.join(self.config.excel_dir, filename)
        try:
            wb = self.excel_manager.load_workbook(full_path)
        except (FileNotFoundError, PermissionError, RuntimeError) as e:
            return None, f"Error: {e}", None, None

        if "Company shortlist" not in wb.sheetnames:
            return None, "Error: sheet 'Company shortlist' not found", None, None
        ws = wb["Company shortlist"]

        try:
            structure = detect_sheet_structure(ws)
        except ValueError as e:
            return None, f"Error detecting sheet structure: {e}", None, None

        return profile_type, full_path, ws, structure

    def describe_sheet(
        self,
        filename: str,
        show_row: Optional[str] = None,
        list_contacts: bool = False,
    ) -> Tuple[Optional[str], Optional[str]]:
        """
        Read-only inspection of the Company shortlist sheet.

        Backs the ``--list-contacts`` / ``--show-row`` helper flags so routine
        verification goes through the script (one allow-listable invocation)
        instead of ad-hoc ``py -c`` one-liners. Never writes the workbook.

        Returns ``(error, output)`` — exactly one is non-None.
        """
        full_path = os.path.join(self.config.excel_dir, filename)
        try:
            wb = self.excel_manager.load_workbook(full_path)
        except (FileNotFoundError, PermissionError, RuntimeError) as e:
            return f"Error: {e}", None
        if "Company shortlist" not in wb.sheetnames:
            return "Error: sheet 'Company shortlist' not found", None
        ws = wb["Company shortlist"]
        try:
            structure = detect_sheet_structure(ws)
        except ValueError as e:
            return f"Error detecting sheet structure: {e}", None

        lines: List[str] = []

        if list_contacts:
            letters = structure.contact_col_letters()
            if not letters:
                lines.append("No key-contact columns found.")
            else:
                lines.append(f"Key contacts ({len(letters)}):")
                for name, col in letters.items():
                    lines.append(f"  - {name} (column {col})")

        if show_row is not None:
            # Locate the company row by case-insensitive name match on the
            # Company column, scanning the data region.
            target = show_row.strip().lower()
            found_row = None
            for r in range(structure.data_start_row, ws.max_row + 1):
                val = ws.cell(row=r, column=structure.company_col).value
                if val and str(val).strip().lower() == target:
                    found_row = r
                    break
            if found_row is None:
                lines.append(f"Company '{show_row}' not found.")
            else:
                lines.append(f"Row {found_row} — {show_row}:")
                # Order mirrors the sheet's own left-to-right layout. Priority and
                # Momentum signal are skipped automatically on a v01-era sheet,
                # where both resolve to None.
                labelled = [
                    ("Company", structure.company_col),
                    ("Priority", structure.priority_col),
                    ("Momentum signal", structure.signal_col),
                    ("Linkedin URL", structure.url_col),
                    ("Industry", structure.industry_col),
                    ("Status", structure.status_col),
                ]
                labelled += list(structure.contact_cols.items())
                for label, col in labelled:
                    if not col:
                        continue
                    cell = ws.cell(row=found_row, column=col).value
                    lines.append(f"  {label}: {cell if cell is not None else ''}")

        return None, "\n".join(lines)

    def plan_urls(
        self,
        name: str,
        linkedin_url: str,
        filename: str,
        max_pages: int = 5,
    ) -> Tuple[Optional[str], List[str]]:
        """
        Enumerate the LinkedIn URLs a run will fetch — no browser, no Excel writes.

        Returns ``(error, urls)``. ``error`` is None on success.
        """
        profile_type, full_path, ws, structure = self._open_sheet(linkedin_url, filename)
        if profile_type is None:
            return full_path, []  # full_path carries the error message
        urls = plan_fetch_urls(
            ws, structure, profile_type, name, linkedin_url,
            parser=self.url_parser, max_pages=max_pages,
        )
        return None, urls

    def execute(
        self,
        name: str,
        linkedin_url: str,
        filename: str,
        refresh: bool = False,
        contacts: Optional[List[Dict[str, str]]] = None,
    ) -> str:
        """
        Run a research pass and write the sheet.

        ``contacts`` (optional) batches several introducers into one run: a list
        of ``{"name", "url"}`` dicts. When given, each company's people page is
        fetched once and filtered for every contact, so N columns cost one set of
        LinkedIn requests instead of N. When omitted the positional
        ``name``/``linkedin_url`` drive a single person- or company-mode run.
        """
        if contacts:
            # Batch person mode — the type is implied; validate via the first URL.
            probe_url = linkedin_url or contacts[0].get("url", "")
            profile_type, full_path, ws, structure = self._open_sheet(probe_url, filename)
            if profile_type is None:
                return full_path
            return self._add_contact_columns(ws, structure, contacts, full_path, refresh)

        profile_type, full_path, ws, structure = self._open_sheet(linkedin_url, filename)
        if profile_type is None:
            return full_path  # carries the error message

        if profile_type == "person":
            return self._add_contact_columns(
                ws, structure, [{"name": name, "url": linkedin_url}], full_path, refresh
            )
        elif profile_type == "company":
            return self._add_company_row(ws, structure, name, linkedin_url, full_path, refresh)
        else:
            return f"Error: unknown profile type '{profile_type}'"

    def init_sheet(
        self,
        filename: str,
        intake: Dict[str, List[Dict[str, str]]],
        overwrite: bool = False,
    ) -> str:
        """
        Create a fresh Company shortlist workbook from the shared schema.

        ``intake`` is gathered conversationally by Claude at first run and holds
        two ordered lists:
          ``contacts``  — ``{"name", "url"}`` introducer columns to seed
          ``companies`` — ``{"company", "url", "industry"}`` starting rows

        The layout matches :data:`SHEET_SCHEMA` exactly, so a created sheet is
        indistinguishable from a hand-maintained one. Refuses to clobber an
        existing file unless ``overwrite`` is set.
        """
        if not filename or not filename.endswith(".xlsx"):
            return "Error: filename must be an .xlsx file"

        full_path = os.path.join(self.config.excel_dir, filename)
        if os.path.exists(full_path) and not overwrite:
            return (
                f"Error: '{full_path}' already exists. "
                "Pass --overwrite to recreate it from scratch."
            )

        contacts = intake.get("contacts", [])
        companies = intake.get("companies", [])
        wb = create_sheet_from_schema(contacts=contacts, companies=companies)

        os.makedirs(self.config.excel_dir, exist_ok=True)
        try:
            wb.save(full_path)
        except PermissionError:
            return (
                f"Error: could not write '{full_path}' — is it open in Excel? "
                "Close it and re-run."
            )

        return (
            f"\n✓ Created '{filename}'\n"
            f"  Contact columns : {len(contacts)}"
            + (f" ({', '.join(c['name'] for c in contacts)})" if contacts else "")
            + f"\n  Seed companies  : {len(companies)}"
            + (f" ({', '.join(c['company'] for c in companies)})" if companies else "")
            + f"\n  Path            : {full_path}"
        )

    # ------------------------------------------------------------------ #
    # Contact-column workflow                                              #
    # ------------------------------------------------------------------ #

    def _add_contact_columns(
        self,
        ws,
        structure: SheetStructure,
        contacts: List[Dict[str, str]],
        file_path: str,
        refresh: bool,
    ) -> str:
        """
        Add/refresh one or more contact columns in a single pass.

        Each company's people page is fetched once (via ``_fetch_company_pages``)
        and filtered for every requested contact, so a batch of N introducers
        costs one set of LinkedIn requests per company rather than N — the whole
        point of WP5's batching. A single contact is just the N = 1 case, so the
        original one-contact flow is preserved.
        """
        backup_path = self.excel_manager.create_backup(file_path)

        # Resolve a target column for every contact up front. Columns are inserted
        # sequentially (each sets its header, so the next ``next_contact_col_idx``
        # finds the following empty slot); afterwards indices are stable.
        targets: List[Tuple[str, int]] = []  # (contact_name, col_idx)
        skipped: List[str] = []              # "name (reason)" for the report
        for contact in contacts:
            contact_name = contact["name"]
            contact_url = contact.get("url")
            existing_col_idx, existing_header = self._find_contact_column(
                ws, structure, contact_name
            )
            if existing_col_idx and not refresh:
                fuzzy_note = (
                    f" (near-match: '{existing_header}')"
                    if existing_header != contact_name else ""
                )
                skipped.append(
                    f"{contact_name}{fuzzy_note} — already in column "
                    f"{get_column_letter(existing_col_idx)}; use --refresh to update"
                )
                continue

            if existing_col_idx:  # refresh existing column in place
                self._clear_column_data(ws, structure, existing_col_idx)
                targets.append((contact_name, existing_col_idx))
            else:                 # brand-new column appended to the right
                new_col_idx = structure.next_contact_col_idx(ws)
                if not new_col_idx:
                    skipped.append(f"{contact_name} — no available column")
                    continue
                self._insert_column_with_header(
                    ws, structure, new_col_idx, contact_name, contact_url
                )
                # Register the freshly inserted column in the live snapshot so a
                # later contact in this same batch (e.g. a repeated name) is
                # deduplicated against it — detection ran before any inserts.
                structure.contact_cols[contact_name] = new_col_idx
                targets.append((contact_name, new_col_idx))

        if not targets:
            # Nothing to do — don't leave a stray backup behind.
            self.excel_manager.delete_backup(backup_path)
            return "No contacts to add:\n  - " + "\n  - ".join(skipped)

        # One row per contact tracks its own (company, count, value) results.
        per_contact: Dict[str, List[Tuple[str, int, str]]] = {n: [] for n, _ in targets}
        company_rows = [
            r for r in range(structure.data_start_row, ws.max_row + 1)
            if ws.cell(row=r, column=structure.company_col).value
        ]
        total = len(company_rows)

        i = 0  # bound before the loop so the interrupt handler can report safely
        try:
            for i, row in enumerate(company_rows):
                company_name = ws.cell(row=row, column=structure.company_col).value
                company_url = (
                    ws.cell(row=row, column=structure.url_col).value
                    if structure.url_col else None
                )

                # Fetch this company's people page ONCE, reuse for all contacts.
                pages = self._fetch_company_pages(str(company_name), company_url)

                if (i + 1) % 5 == 0:
                    print(f"  Progress: {i+1}/{total} companies")

                for contact_name, col_idx in targets:
                    connections = self._connections_for_contact(
                        pages, str(company_name), contact_name
                    )
                    cell_value = ", ".join(connections) if connections else ""
                    ws.cell(row=row, column=col_idx).value = cell_value or None
                    per_contact[contact_name].append(
                        (company_name, len(connections), cell_value)
                    )

        except KeyboardInterrupt:
            print("\n[Interrupted] Saving partial results...")
            self._refresh_filters(ws)
            self._save_or_restore(file_path, backup_path)
            return f"Interrupted. Partial results saved ({i} of {total} companies processed)."

        self._refresh_filters(ws)
        if not self._save_or_restore(file_path, backup_path):
            return "Error: Failed to save workbook. Original restored from backup."

        return self._contacts_report(per_contact, skipped)

    # ------------------------------------------------------------------ #
    # Company-row workflow                                                 #
    # ------------------------------------------------------------------ #

    def _add_company_row(
        self,
        ws,
        structure: SheetStructure,
        company_name: str,
        linkedin_url: str,
        file_path: str,
        refresh: bool,
    ) -> str:
        if not refresh:
            existing_row, existing_name = self._find_company_row(ws, structure, company_name)
            if existing_row:
                fuzzy_note = (
                    f" (near-match: '{existing_name}')" if existing_name != company_name else ""
                )
                return (
                    f"Company '{company_name}'{fuzzy_note} already exists in row {existing_row}. "
                    f"Run with --refresh to update."
                )

        backup_path = self.excel_manager.create_backup(file_path)

        # Fetch company profile (industry) — requires browser
        industry = "Unknown (verify)"
        if self.researcher.fetch_fn:
            profile = self.researcher.fetch_company_profile(linkedin_url)
            if profile:
                industry = profile.get("industry", "Unknown (verify)")
                if industry == "Unknown (verify)":
                    print(f"[Warning] Industry not found on LinkedIn — please verify manually")

        if refresh:
            target_row, _ = self._find_company_row(ws, structure, company_name)
            if not target_row:
                target_row = self._find_next_row(ws, structure)
                self._insert_company_row(ws, structure, target_row, company_name, linkedin_url, industry)
            else:
                # Overwrite company info, clear contact cells
                if structure.url_col:
                    ws.cell(row=target_row, column=structure.url_col).value = linkedin_url
                if structure.industry_col:
                    ws.cell(row=target_row, column=structure.industry_col).value = industry
                for col_idx in structure.contact_cols.values():
                    ws.cell(row=target_row, column=col_idx).value = None
        else:
            target_row = self._find_next_row(ws, structure)
            self._insert_company_row(ws, structure, target_row, company_name, linkedin_url, industry)

        # Research all key contacts
        results: List[Tuple[str, bool, str]] = []
        total_conns = 0
        contact_items = list(structure.contact_cols.items())

        try:
            for contact_name, col_idx in contact_items:
                print(f"  Researching {contact_name} at {company_name}...")
                connections = self._research_connections(
                    company_name, linkedin_url, contact_name
                )
                total_conns += len(connections)
                cell_value = ", ".join(connections) if connections else ""
                ws.cell(row=target_row, column=col_idx).value = cell_value or None
                results.append((contact_name, bool(connections), cell_value))

        except KeyboardInterrupt:
            print("\n[Interrupted] Saving partial results...")
            self._refresh_filters(ws)
            self._save_or_restore(file_path, backup_path)
            return f"Interrupted. Partial results saved ({len(results)} of {len(contact_items)} contacts)."

        self._refresh_filters(ws)
        if not self._save_or_restore(file_path, backup_path):
            return "Error: Failed to save workbook. Original restored from backup."

        found = sum(1 for _, f, _ in results if f)
        return self._company_report(company_name, industry, results, found, total_conns)

    # ------------------------------------------------------------------ #
    # Internal helpers                                                     #
    # ------------------------------------------------------------------ #

    def _research_connections(
        self,
        company_name: str,
        company_url: Optional[str],
        contact_name: str,
    ) -> List[str]:
        """Return formatted connection strings. Empty list if browser not wired."""
        if not self.researcher.fetch_fn:
            return []
        search_url = self.url_parser.build_people_search_url(
            company_name=company_name,
            company_url=company_url,
        )
        return self.researcher.find_connections(
            company_name=company_name,
            contact_name=contact_name,
            role_patterns=self.config.role_patterns(),
            search_url=search_url,
        )

    def _fetch_company_pages(
        self, company_name: str, company_url: Optional[str]
    ) -> List[str]:
        """
        Fetch one company's people-page text (contact-independent).

        The search URL depends only on the company, so this is fetched once per
        company and reused across every contact in a batch. Returns [] when no
        browser is wired (offline dry-run).
        """
        if not self.researcher.fetch_fn:
            return []
        search_url = self.url_parser.build_people_search_url(
            company_name=company_name,
            company_url=company_url,
        )
        # Pass the expected company so the researcher can skip (with a warning) a
        # people tab that resolved to the wrong org / an "unavailable" redirect.
        return self.researcher.search_company_employees(
            search_url, expected_company=company_name
        )

    def _connections_for_contact(
        self, pages: List[str], company_name: str, contact_name: str
    ) -> List[str]:
        """Filter already-fetched company pages for one contact's connections."""
        if not pages or self.researcher.role_matcher is None:
            return []
        connections = find_connections_in_pages(
            pages, contact_name, self.researcher.role_matcher, default_company=company_name
        )
        return [c.format_for_excel() for c in connections]

    def _insert_column_with_header(
        self,
        ws,
        structure: SheetStructure,
        target_col_idx: int,
        contact_name: str,
        url: Optional[str] = None,
    ) -> None:
        """
        Insert a column at target_col_idx, copy formatting from the left neighbour.

        Writes the contact's full name as the header and, when ``url`` is given,
        the contact's own profile URL on ``contact_url_row`` (row 5) — mirroring
        the init-sheet path so appended columns match created ones.
        """
        # Capture source style BEFORE inserting (the left neighbour)
        source_col_idx = target_col_idx - 1
        source_styles: Dict[int, object] = {}
        for row in range(structure.header_row, ws.max_row + 1):
            src = ws.cell(row=row, column=source_col_idx)
            if src.has_style:
                source_styles[row] = {
                    "font": copy.copy(src.font),
                    "border": copy.copy(src.border),
                    "fill": copy.copy(src.fill),
                    "number_format": src.number_format,
                    "alignment": copy.copy(src.alignment),
                    "protection": copy.copy(src.protection),
                }
        source_width = ws.column_dimensions[
            get_column_letter(source_col_idx)
        ].width

        # Insert the new column
        ws.insert_cols(target_col_idx)

        # Apply captured styles
        for row, style in source_styles.items():
            tgt = ws.cell(row=row, column=target_col_idx)
            tgt.font = style["font"]
            tgt.border = style["border"]
            tgt.fill = style["fill"]
            tgt.number_format = style["number_format"]
            tgt.alignment = style["alignment"]
            tgt.protection = style["protection"]
        if source_width:
            ws.column_dimensions[get_column_letter(target_col_idx)].width = source_width

        # Set header to the contact's (full) name — the concise, load-bearing key.
        ws.cell(row=structure.header_row, column=target_col_idx).value = contact_name

        # Write the contact's own profile URL beneath the header (row 5), matching
        # the init-sheet path. Only new columns write this; --refresh leaves it be.
        if url:
            ws.cell(row=structure.contact_url_row, column=target_col_idx).value = url

    def _insert_company_row(
        self,
        ws,
        structure: SheetStructure,
        target_row: int,
        company_name: str,
        linkedin_url: str,
        industry: str,
    ) -> None:
        """Insert a new row at target_row and populate company info."""
        # Capture source style BEFORE inserting
        source_row = target_row - 1
        max_col = ws.max_column
        source_styles: Dict[int, dict] = {}
        for col in range(1, max_col + 1):
            src = ws.cell(row=source_row, column=col)
            if src.has_style:
                source_styles[col] = {
                    "font": copy.copy(src.font),
                    "border": copy.copy(src.border),
                    "fill": copy.copy(src.fill),
                    "number_format": src.number_format,
                    "alignment": copy.copy(src.alignment),
                    "protection": copy.copy(src.protection),
                }

        ws.insert_rows(target_row)

        # Apply captured styles
        for col, style in source_styles.items():
            tgt = ws.cell(row=target_row, column=col)
            tgt.font = style["font"]
            tgt.border = style["border"]
            tgt.fill = style["fill"]
            tgt.number_format = style["number_format"]
            tgt.alignment = style["alignment"]
            tgt.protection = style["protection"]

        # Write company data
        ws.cell(row=target_row, column=structure.company_col).value = company_name
        if structure.url_col:
            ws.cell(row=target_row, column=structure.url_col).value = linkedin_url
        if structure.industry_col:
            ws.cell(row=target_row, column=structure.industry_col).value = industry
        if structure.status_col:
            ws.cell(row=target_row, column=structure.status_col).value = self.DEFAULT_STATUS
        if structure.priority_col:
            ws.cell(row=target_row, column=structure.priority_col).value = self.DEFAULT_PRIORITY
        if structure.signal_col:
            # Blank but present: the rating's justification is written by hand, and an
            # empty cell here is the visible cue that it is still owed. Priority and
            # signal always travel together.
            ws.cell(row=target_row, column=structure.signal_col).value = None

    def _clear_column_data(self, ws, structure: SheetStructure, col_idx: int) -> None:
        """Clear data cells (not header) in a contact column."""
        for row in range(structure.data_start_row, ws.max_row + 1):
            ws.cell(row=row, column=col_idx).value = None

    def _find_contact_column(
        self, ws, structure: SheetStructure, contact_name: str
    ) -> Tuple[Optional[int], Optional[str]]:
        """
        Find if a contact column already exists (headers are bare names).

        Scans only the detected contact columns — never the fixed columns — so a
        bare name can't collide with 'Company'/'Industry'/etc. Match order:
        exact (case-insensitive) → first-name equivalence (a single-token name
        matches the other side's first token, so the migration's first-name-only
        headers like "Wren" match an "Wren Rivera" lookup) → length-guarded
        containment → difflib fuzzy fallback for near-dupes.
        Returns (col_idx, actual_header_text) or (None, None).
        """
        contact_lower = contact_name.lower().strip()
        contact_tokens = contact_lower.split()
        best_ratio = 0.0
        best_col: Optional[int] = None
        best_header: Optional[str] = None

        for header, col_idx in structure.contact_cols.items():
            header_lower = str(header).lower().strip()
            header_tokens = header_lower.split()
            if contact_lower == header_lower:
                return col_idx, header
            # First-name equivalence: when exactly one side is a single token,
            # treat it as the same contact if it equals the other's first token.
            # Bridges the migration's first-name-only headers ("Wren") against a
            # full-name lookup ("Wren Rivera"); two distinct full names sharing a
            # first name are both multi-token, so this branch is skipped for them.
            if (
                (len(contact_tokens) == 1) != (len(header_tokens) == 1)
                and contact_tokens and header_tokens
                and contact_tokens[0] == header_tokens[0]
            ):
                return col_idx, header
            # Containment catches longer overlaps like "Toni Vale" vs
            # "Toni Vale-Hollis", but only when both names are long enough that the
            # overlap is meaningful (mirrors the company matcher's guard).
            if min(len(contact_lower), len(header_lower)) >= 5 and (
                contact_lower in header_lower or header_lower in contact_lower
            ):
                return col_idx, header
            ratio = difflib.SequenceMatcher(None, contact_lower, header_lower).ratio()
            if ratio > best_ratio:
                best_ratio = ratio
                best_col = col_idx
                best_header = header

        if best_ratio >= self.FUZZY_MATCH_THRESHOLD:
            return best_col, best_header
        return None, None

    def _find_company_row(
        self, ws, structure: SheetStructure, company_name: str
    ) -> Tuple[Optional[int], Optional[str]]:
        """
        Find if a company row already exists.
        Returns (row_idx, actual_name) or (None, None).
        """
        name_lower = company_name.lower()
        best_ratio = 0.0
        best_row: Optional[int] = None
        best_name: Optional[str] = None

        for row in range(structure.data_start_row, ws.max_row + 1):
            val = ws.cell(row=row, column=structure.company_col).value
            if not val:
                continue
            val_str = str(val)
            if name_lower == val_str.lower():
                return row, val_str
            # Substring check — catches "Navan" vs "Navan Inc." — but only when the
            # contained name is long enough for the overlap to be meaningful. Without
            # this length guard a short name like "Meta" would spuriously match
            # "Metabase" (and "AWS" any row containing "aws").
            if min(len(name_lower), len(val_str)) >= 5 and (
                name_lower in val_str.lower() or val_str.lower() in name_lower
            ):
                return row, val_str
            ratio = difflib.SequenceMatcher(None, name_lower, val_str.lower()).ratio()
            if ratio > best_ratio:
                best_ratio = ratio
                best_row = row
                best_name = val_str

        if best_ratio >= self.FUZZY_MATCH_THRESHOLD:
            return best_row, best_name
        return None, None

    def _find_next_row(self, ws, structure: SheetStructure) -> int:
        """Find the first empty company row after the last data row."""
        last_data_row = structure.data_start_row
        for row in range(ws.max_row, structure.data_start_row - 1, -1):
            if ws.cell(row=row, column=structure.company_col).value:
                last_data_row = row
                break
        return last_data_row + 1

    def _refresh_filters(self, ws) -> None:
        """Re-span the autofilter before saving.

        Structural changes (a company row appended, a contact column added) leave
        the stored filter range behind, so rows and columns added after the filter
        was first applied drop out of filtering and sorting. Re-detecting rather
        than reusing the caller's ``structure`` keeps this correct when a contact
        column was added during the run, which the caller's copy predates.
        """
        try:
            structure = detect_sheet_structure(ws)
        except ValueError:
            return  # unreadable layout — saving the data still matters more
        refresh_autofilter(ws, structure)

    def _save_or_restore(self, file_path: str, backup_path: str) -> bool:
        """
        Save workbook; restore backup if save fails. Returns True on success.

        On a confirmed successful save the crash-safety backup is deleted (it is
        no longer needed and must not accumulate). On failure the original is
        restored from the backup, which is kept so the user can retry.
        """
        try:
            self.excel_manager.save_workbook(file_path)
            self.excel_manager.delete_backup(backup_path)
            self.last_save_ok = True
            return True
        except Exception as e:
            print(f"[Warning] Save failed ({e}). Restoring backup...")
            self.excel_manager.restore_backup(backup_path, file_path)
            self.last_save_ok = False
            return False

    # ------------------------------------------------------------------ #
    # Reports                                                              #
    # ------------------------------------------------------------------ #

    def _contacts_report(
        self,
        per_contact: Dict[str, List[Tuple[str, int, str]]],
        skipped: List[str],
    ) -> str:
        """Render a per-contact summary for a (possibly batched) contact run."""
        lines: List[str] = []
        for contact_name, results in per_contact.items():
            found = sum(1 for _, count, _ in results if count)
            total_conns = sum(count for _, count, _ in results)
            lines += [
                f"\n✓ Contact column added for '{contact_name}'",
                f"  Companies researched : {len(results)}",
                f"  Companies with hits  : {found}",
                f"  Total connections    : {total_conns}",
                "",
            ]
            for company, count, conn_str in results:
                if count:
                    lines.append(f"  • {company}: {conn_str}")

        if skipped:
            lines += ["", "  Skipped:"]
            lines += [f"    - {s}" for s in skipped]

        lines += ["", f"  Saved : {datetime.now():%Y-%m-%d %H:%M:%S}"]
        return "\n".join(lines)

    def _company_report(
        self,
        company_name: str,
        industry: str,
        results: List[Tuple],
        found: int,
        total_conns: int,
    ) -> str:
        lines = [
            f"\n✓ Company row added for '{company_name}'",
            f"  Industry             : {industry}",
            f"  Contacts researched  : {len(results)}",
            f"  Contacts with hits   : {found}",
            f"  Total connections    : {total_conns}",
            "",
        ]
        for contact, hit, conn_str in results:
            if hit:
                lines.append(f"  • {contact}: {conn_str}")
        lines += ["", f"  Saved : {datetime.now():%Y-%m-%d %H:%M:%S}"]
        return "\n".join(lines)


# --------------------------------------------------------------------------- #
# CLI entry point                                                              #
# --------------------------------------------------------------------------- #

def main():
    import argparse

    # Reports contain Unicode glyphs (✓, •). The default Windows console codec
    # (cp1252) can't encode them and print() would crash, so force UTF-8 output.
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except (AttributeError, ValueError):
        pass

    parser = argparse.ArgumentParser(description="Find Connections (Basic) skill")
    # Positionals are optional so --init-sheet (which takes no name/url) can share
    # the parser; the research modes validate their presence below.
    parser.add_argument("name", nargs="?", default=None, help="Contact or company name")
    parser.add_argument("url", nargs="?", default=None, help="LinkedIn profile URL")
    parser.add_argument(
        "filename", nargs="?", default=None,
        help="Excel filename (defaults to excel_filename from the per-user config)",
    )
    parser.add_argument("--refresh", action="store_true", help="Overwrite existing data")

    # Claude-in-Chrome bridge modes (see SKILL.md). Mutually exclusive with a
    # plain offline run. Claude orchestrates: --plan → fetch pages → --apply.
    mode = parser.add_mutually_exclusive_group()
    mode.add_argument(
        "--plan", action="store_true",
        help="Print the list of LinkedIn URLs to fetch (no browser, no Excel writes)",
    )
    mode.add_argument(
        "--apply", action="store_true",
        help="Run using a Chrome-fetched response cache and write Excel",
    )
    mode.add_argument(
        "--init-sheet", action="store_true",
        help="Create a fresh Company shortlist workbook from an --intake JSON",
    )
    parser.add_argument(
        "--intake",
        help="Path to the first-run intake JSON ({contacts:[...], companies:[...]})",
    )
    parser.add_argument(
        "--overwrite", action="store_true",
        help="With --init-sheet/--setup, recreate the workbook even if it already exists",
    )
    parser.add_argument(
        "--setup", metavar="SETUP_JSON",
        help="First-run wizard: write the per-user config AND create a starter "
             "sheet in one shot, from a JSON of {config:{...}, intake:{...}}",
    )
    parser.add_argument(
        "--contacts",
        help="Path to a JSON list [{name, url}, ...] to add several contact "
             "columns in one batched run (person mode)",
    )
    parser.add_argument(
        "--cache",
        help="Path to the {url: page_text} response cache JSON "
             "(default: ephemeral work dir)",
    )
    parser.add_argument("--max-pages", type=int, default=5, help="Search result pages per company")

    # Read-only / work-dir-only helper actions. These exist so routine discovery
    # and verification go through this script (a single allow-listable command)
    # rather than ad-hoc `py -c` one-liners. Each handler prints and exits.
    parser.add_argument(
        "--print-cache-path", action="store_true",
        help="Print the response-cache path (honoring --cache) and exit",
    )
    parser.add_argument(
        "--build-cache", metavar="MAP_JSON",
        help='Assemble the response cache from a JSON map {"<url>": "<page_text_file>"}; '
             "writes to --cache or the default work-dir cache, then exits",
    )
    parser.add_argument(
        "--show-row", metavar="COMPANY",
        help="Print the Company shortlist row for COMPANY (read-only) and exit",
    )
    parser.add_argument(
        "--list-contacts", action="store_true",
        help="Print the sheet's key-contact columns (read-only) and exit",
    )
    args = parser.parse_args()

    # The workbook is the THIRD positional (name, url, filename). A bare
    # "script.py --list-contacts book.xlsx" puts the filename in `name`, leaving
    # `filename` None — the run then silently falls back to the configured default
    # and reports the wrong workbook's layout. Refuse rather than guess: this has
    # already produced results attributed to the wrong file.
    _misplaced = [
        (slot, value)
        for slot, value in (("name", args.name), ("url", args.url))
        if isinstance(value, str) and value.lower().endswith((".xlsx", ".xlsm"))
    ]
    if _misplaced and not args.filename:
        slot, value = _misplaced[0]
        print(
            "Error: '" + value + "' looks like a workbook but was passed "
            "as the '" + slot + "' argument." + "\n"
            "       The workbook is the third positional argument:" + "\n"
            "           script.py NAME URL " + value + "\n"
            "       For the sheet-inspection flags pass empty name/url first, "
            "or set 'excel_filename' in config.json instead."
        )
        # Non-zero: this guard exists to stop a run that would otherwise operate on
        # the wrong workbook, so a caller must not read it as a successful no-op.
        sys.exit(1)

    # --- First-run wizard -----------------------------------------------------
    # --setup writes the per-user config AND a starter sheet in one shot. It is
    # handled BEFORE UserConfig.load() because on a genuine first run no config
    # exists yet — this is the command that creates it.
    if args.setup:
        import json

        try:
            with open(args.setup, encoding="utf-8") as f:
                setup = json.load(f)
        except (OSError, ValueError) as e:
            print(f"Error reading --setup JSON '{args.setup}': {e}")
            return
        if not isinstance(setup, dict):
            print('Error: --setup JSON must be an object with "config" (and optional "intake") keys.')
            return
        cfg_data = setup.get("config") or {}
        missing = [k for k in ("excel_dir", "excel_filename") if not cfg_data.get(k)]
        if missing:
            print(f"Error: --setup config is missing required keys: {missing}")
            return
        intake = setup.get("intake") or {"contacts": [], "companies": []}
        if not all(isinstance(c, dict) and c.get("name") for c in intake.get("contacts", [])):
            print('Error: each --setup intake contact needs a "name" (and optional "url").')
            return
        if not all(isinstance(c, dict) and c.get("company") for c in intake.get("companies", [])):
            print('Error: each --setup intake company needs a "company".')
            return

        new_config = UserConfig(
            excel_dir=cfg_data["excel_dir"],
            excel_filename=cfg_data["excel_filename"],
            relevant_roles=cfg_data.get("relevant_roles", []),
            relevant_industries=cfg_data.get("relevant_industries", []),
            relevant_role_patterns=cfg_data.get("relevant_role_patterns", []),
        )
        config_path = new_config.save()
        skill = LinkedInExcelContactSkill(new_config)
        sheet_filename = args.filename or new_config.excel_filename
        sheet_msg = skill.init_sheet(sheet_filename, intake, overwrite=args.overwrite)
        # init_sheet returns either a "\n✓ Created …" success block or an
        # "Error: …" string (e.g. the sheet already exists). The config was saved
        # either way, so don't run a ✓ line straight into an error — report the
        # sheet outcome separately and point at --overwrite.
        if sheet_msg.lstrip().startswith("Error"):
            print(f"\n✓ Wrote per-user config → {config_path}")
            print(f"⚠ Sheet not created — {sheet_msg.strip()}")
        else:
            print(f"\n✓ Wrote per-user config → {config_path}{sheet_msg}")
        return

    # The per-user config supplies the data location and role criteria. If it is
    # missing this is a first run: tell the orchestrator to create it (Claude
    # gathers the values in chat and writes the file) rather than guessing.
    try:
        config = UserConfig.load()
    except ConfigNotFoundError:
        print(
            "First-run setup needed: no per-user config found at\n"
            f"  {UserConfig.config_path()}\n"
            "Run the wizard with --setup <setup.json> (a JSON of {config:{...}, "
            "intake:{...}}) to write the config and a starter sheet in one shot;\n"
            "see scripts/config.example.json for the config keys."
        )
        return

    filename = args.filename or config.excel_filename

    # --- Helper actions (print-and-exit) -------------------------------------
    # Kept ahead of the name/url validation because none of them need a
    # positional contact/company — they replace the ad-hoc `py -c` snippets.
    if args.print_cache_path:
        print(args.cache or default_cache_path())
        return

    if args.build_cache:
        import json

        try:
            with open(args.build_cache, encoding="utf-8") as f:
                url_to_file = json.load(f)
        except (OSError, ValueError) as e:
            print(f"Error reading --build-cache map '{args.build_cache}': {e}")
            return
        if (
            not isinstance(url_to_file, dict)
            or not url_to_file
            or not all(isinstance(v, str) for v in url_to_file.values())
        ):
            print(
                'Error: --build-cache JSON must be a non-empty {"<url>": "<page_text_file>"} '
                "map with string file-path values."
            )
            return
        cache: Dict[str, str] = {}
        for url, text_file in url_to_file.items():
            try:
                with open(text_file, encoding="utf-8") as f:
                    cache[url] = f.read()
            except OSError as e:
                print(f"Error reading page-text file '{text_file}' for {url}: {e}")
                sys.exit(1)
        cache_path = args.cache or default_cache_path()
        try:
            # dirname is "" when --cache is a bare filename; skip makedirs then.
            cache_dir = os.path.dirname(cache_path)
            if cache_dir:
                os.makedirs(cache_dir, exist_ok=True)
            with open(cache_path, "w", encoding="utf-8") as f:
                json.dump(cache, f)
        except OSError as e:
            print(f"Error writing cache to '{cache_path}': {e}")
            sys.exit(1)
        plural = "entry" if len(cache) == 1 else "entries"
        print(f"Wrote {len(cache)} cache {plural} to {cache_path}")
        return

    if args.list_contacts or args.show_row is not None:
        skill = LinkedInExcelContactSkill(config)
        error, output = skill.describe_sheet(
            filename, show_row=args.show_row, list_contacts=args.list_contacts
        )
        print(error or output)
        return

    if args.init_sheet:
        import json

        intake: Dict[str, List[Dict[str, str]]] = {"contacts": [], "companies": []}
        if args.intake:
            try:
                with open(args.intake, encoding="utf-8") as f:
                    intake = json.load(f)
            except (OSError, ValueError) as e:
                print(f"Error reading intake JSON '{args.intake}': {e}")
                return
            if not all(
                isinstance(c, dict) and c.get("name")
                for c in intake.get("contacts", [])
            ):
                print('Error: each --intake contact needs a "name" (and optional "url").')
                return
            if not all(
                isinstance(c, dict) and c.get("company")
                for c in intake.get("companies", [])
            ):
                print(
                    'Error: each --intake company needs a "company" '
                    '(and optional "url"/"industry").'
                )
                return
        skill = LinkedInExcelContactSkill(config)
        print(skill.init_sheet(filename, intake, overwrite=args.overwrite))
        return

    # Optional batch of contacts (person mode). When given, positional name/url
    # are not required — the URL list is contact-independent anyway.
    contacts: Optional[List[Dict[str, str]]] = None
    if args.contacts:
        import json

        try:
            with open(args.contacts, encoding="utf-8") as f:
                contacts = json.load(f)
        except (OSError, ValueError) as e:
            print(f"Error reading contacts JSON '{args.contacts}': {e}")
            return
        if not contacts:
            print("Error: --contacts JSON is empty.")
            return
        if not isinstance(contacts, list) or not all(
            isinstance(c, dict) and c.get("name") and c.get("url") for c in contacts
        ):
            print('Error: --contacts JSON must be a list of {"name", "url"} objects.')
            return

    # Research modes need a name + URL (or a --contacts batch); init-sheet does not.
    if not contacts and (not args.name or not args.url):
        print("Error: name and LinkedIn URL are required (unless using --init-sheet or --contacts).")
        return

    # For batch runs the URL enumeration is company-driven, so any person URL
    # anchors profile-type detection; use the first contact's.
    probe_name = args.name or (contacts[0]["name"] if contacts else None)
    probe_url = args.url or (contacts[0]["url"] if contacts else None)

    if args.plan:
        skill = LinkedInExcelContactSkill(config)  # no browser needed to plan
        error, urls = skill.plan_urls(probe_name, probe_url, filename, args.max_pages)
        if error:
            print(error)
            return
        # No queue file: the URL list is the stdout below. Claude reads it here,
        # drives Chrome, and writes the response cache to --cache (work dir).
        print(f"Planned {len(urls)} URL(s) to fetch:")
        for u in urls:
            print(f"  - {u}")
        return

    if args.apply:
        cache_path = args.cache or default_cache_path()
        fetcher = CachedFetcher(cache_path)
        skill = LinkedInExcelContactSkill(config, fetch_fn=fetcher)
        skill.cache_path = cache_path
        skill.researcher.min_delay = 0.0  # cached lookups need no rate-limit delay
        result = skill.execute(probe_name, probe_url, filename, args.refresh, contacts=contacts)
        print(result)
        if fetcher.misses:
            print(
                f"\n[Warning] {len(fetcher.misses)} URL(s) were not in the cache and "
                f"returned empty. Fetch these and re-run --apply:"
            )
            for u in fetcher.misses:
                print(f"  - {u}")
        # Cache is disposable once the run is complete: delete it only when the
        # save succeeded AND nothing was missing (misses mean a re-run needs it).
        elif skill.last_save_ok and os.path.exists(cache_path):
            try:
                os.remove(cache_path)
            except OSError:
                pass  # cache is disposable; a locked/vanished file is harmless
        return

    skill = LinkedInExcelContactSkill(config)  # no fetch_fn → browser calls skipped
    result = skill.execute(probe_name, probe_url, filename, args.refresh, contacts=contacts)
    print(result)


if __name__ == "__main__":
    main()
