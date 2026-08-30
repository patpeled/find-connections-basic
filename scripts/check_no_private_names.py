"""
Pre-publish check: no real person's name from your own spreadsheets appears in this
repository.

This skill's tests and docs need example people. It is very easy for those examples to
end up being your actual contacts - and this repository may be public. Run this before
pushing.

It reads your private workbooks (never modifies them), builds the set of real names and
LinkedIn slugs they contain, and searches this repository for any of them. Nothing it
reads is written anywhere; only a pass/fail summary is printed.

    py scripts/check_no_private_names.py --crm "C:/path/My CRM.xlsx"
    py scripts/check_no_private_names.py --crm crm.xlsx --shortlist other.xlsx --repo .

--shortlist defaults to the workbook named in config.json. --crm has no default: it is
your file and the script will not guess at a path.

Exit status is 0 when clean, 1 when something was found, 2 when the check could not be
trusted (see "Why this refuses to guess" at the bottom).
"""
import argparse
import io
import json
import pathlib
import re
import sys

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required: py -m pip install openpyxl")

SKILL_ROOT = pathlib.Path(__file__).resolve().parent.parent


def load_config():
    p = SKILL_ROOT / "config.json"
    if not p.exists():
        return {}
    try:
        return json.loads(io.open(p, encoding="utf-8").read())
    except (ValueError, OSError):
        return {}


def expand(p):
    import os
    return pathlib.Path(os.path.expandvars(os.path.expanduser(str(p))))


def names_from_crm(path, sheet, first_col, last_col, header_row):
    """First/last name columns of a CRM-style sheet, plus their combination."""
    out = set()
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        # Never fall back to the first sheet. Reading the wrong one produces a result
        # that looks like a real answer - the whole failure mode this tool guards against.
        wb.close()
        raise SystemExit(
            "sheet {!r} not found in {}. Available: {}".format(
                sheet, pathlib.Path(path).name, ", ".join(wb.sheetnames)))
    ws = wb[sheet]
    for r in range(header_row + 1, ws.max_row + 1):
        first = ws.cell(r, first_col).value
        last = ws.cell(r, last_col).value
        for v in (first, last):
            if isinstance(v, str) and len(v.strip()) > 2:
                out.add(v.strip())
        if isinstance(first, str) and isinstance(last, str):
            out.add(first.strip() + " " + last.strip())
    wb.close()
    return out


def names_from_shortlist(path, sheet, header_row):
    """Introducer column headers, 'Name (Role at Company)' cells, and profile slugs."""
    names, slugs = set(), set()
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        raise SystemExit(
            "sheet {!r} not found in {}. Available: {}".format(
                sheet, pathlib.Path(path).name, ", ".join(wb.sheetnames)))
    ws = wb[sheet]
    fixed = {"company", "priority", "linkedin url", "industry", "status"}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(header_row, c).value
        if isinstance(h, str) and h.strip():
            low = h.strip().lower()
            if low not in fixed and not low.startswith("momentum"):
                names.add(h.strip())          # an introducer column is a real person
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if not isinstance(v, str):
                continue
            for m in re.findall(r"([A-Z][a-zA-Z'.-]+(?: [A-Z][a-zA-Z'.-]+){1,2})\s*\(", v):
                names.add(m.strip())
            for m in re.findall(r"linkedin\.com/in/([A-Za-z0-9._-]+)", v):
                slugs.add(m.lower())
    wb.close()
    return names, slugs


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--crm", help="workbook of contacts (first/last name columns)")
    ap.add_argument("--crm-sheet", default="02_People overview")
    ap.add_argument("--crm-first-col", type=int, default=3)
    ap.add_argument("--crm-last-col", type=int, default=4)
    ap.add_argument("--crm-header-row", type=int, default=6)
    ap.add_argument("--shortlist", action="append", default=[],
                    help="shortlist workbook; repeatable. Defaults to config.json's.")
    ap.add_argument("--sheet", default="Company shortlist")
    ap.add_argument("--header-row", type=int, default=4)
    ap.add_argument("--repo", default=str(SKILL_ROOT), help="directory to search")
    ap.add_argument("--min-identities", type=int, default=10,
                    help="refuse to report clean below this many names (default 10)")
    ap.add_argument("--allow", action="append", default=[],
                    help="literal string to neutralise before matching, e.g. a city "
                         "name that collides with a contact's first name")
    args = ap.parse_args()

    shortlists = [expand(s) for s in args.shortlist]
    if not shortlists:
        cfg = load_config()
        if cfg.get("excel_dir") and cfg.get("excel_filename"):
            shortlists = [expand(cfg["excel_dir"]) / cfg["excel_filename"]]

    names, slugs = set(), set()
    for path in shortlists:
        if not path.exists():
            print("  note: {} not found, skipping".format(path.name))
            continue
        n, s = names_from_shortlist(path, args.sheet, args.header_row)
        names |= n
        slugs |= s
        print("  read shortlist: {}".format(path.name))
    if args.crm:
        crm = expand(args.crm)
        if not crm.exists():
            sys.exit("--crm not found: {}".format(crm))
        names |= names_from_crm(crm, args.crm_sheet, args.crm_first_col,
                                args.crm_last_col, args.crm_header_row)
        print("  read CRM: {}".format(crm.name))

    names = {n for n in names if len(n) > 2}
    print("real names: {}   real slugs: {}".format(len(names), len(slugs)))

    # Refuse to report clean off an empty corpus. If a workbook moves or a sheet is
    # renamed, extraction quietly returns nothing, every search below trivially finds
    # no match, and the run reports CLEAN - a false all-clear that looks exactly like
    # a real pass.
    if len(names) < args.min_identities:
        print("\nABORT: only {} names loaded (expected at least {}). A workbook path or "
              "sheet name is probably wrong. Fix that before trusting any result here."
              .format(len(names), args.min_identities))
        return 2

    repo = expand(args.repo)
    corpus = {}
    for f in list(repo.rglob("*.py")) + list(repo.rglob("*.md")):
        if ".git" in f.parts:
            continue
        try:
            corpus[str(f.relative_to(repo))] = io.open(f, encoding="utf-8").read()
        except (OSError, UnicodeDecodeError):
            continue
    for phrase in args.allow:
        for k in corpus:
            corpus[k] = corpus[k].replace(phrase, "_" * len(phrase))
    print("files searched: {}".format(len(corpus)))
    if not corpus:
        print("\nABORT: no files searched under {}.".format(repo))
        return 2

    leaks = []
    for n in sorted(names):
        hits = [f for f, t in corpus.items() if re.search(r"\b" + re.escape(n) + r"\b", t)]
        if hits:
            leaks.append(("name", n, hits))
    for s in sorted(slugs):
        hits = [f for f, t in corpus.items() if s in t.lower()]
        if hits:
            leaks.append(("slug", s, hits))

    print()
    if leaks:
        print("FOUND {} real identit{} in this repository:"
              .format(len(leaks), "y" if len(leaks) == 1 else "ies"))
        for kind, val, hits in leaks:
            print("  {} {!r} in {}".format(kind, val, ", ".join(hits)))
        print("\nReplace these with invented names before publishing.")
        return 1
    print("CLEAN: no name or slug from your workbooks appears in this repository.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
