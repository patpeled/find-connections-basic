"""
Excel utilities for managing the Company Shortlist spreadsheet.
"""

import copy
import os
import shutil
from dataclasses import dataclass, field
from datetime import datetime
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from workspace import work_dir


@dataclass(frozen=True)
class SheetSchema:
    """Single source of truth for the Company shortlist layout.

    Both :func:`detect_sheet_structure` (reader) and
    :func:`create_sheet_from_schema` (writer) consume this, so a fresh sheet is
    laid out exactly like the reference sheet maintained by hand:
    column A is intentionally blank, the title sits at B2, the header row is 4
    (starting at column B), row 5 holds each contact's own LinkedIn URL, and
    company data begins at row 6.

    :meth:`fixed_headers` is the one place the left block's order is declared.
    Anything that writes a fixed column must resolve its position through that
    order (or through the reader) rather than by offset arithmetic.

    The full contract - the layout, the eight rules, and why they are what they
    are - is in ``SHEET_CONTRACT.md`` at the skill root.
    """

    sheet_name: str = "Company shortlist"
    title_cell: str = "B2"
    title_text: str = "Overview of relevant companies and contacts"
    first_col: int = 2  # column B — column A is left empty by design
    header_row: int = 4
    contact_url_row: int = 5  # contacts' own profile URLs live here
    data_start_row: int = 6
    company_header: str = "Company"
    url_header: str = "Linkedin URL"
    industry_header: str = "Industry"
    status_header: str = "Status"
    # Priority and Momentum signal sit *inside* the fixed left block, never at the
    # right end: the contact block grows rightward, so a fixed column appended there
    # would be overwritten by the next contact added.
    priority_header: str = "Priority"
    signal_header: str = "Momentum signal"
    default_status: str = "No successful contact so far"
    # New companies are rated P2 by default — never P1, so nothing is treated as a top
    # target before a human has reviewed it. The rating itself stays a manual act.
    default_priority: str = "P2"
    contact_col_width: float = 45.0
    # Fixed-column widths in :meth:`fixed_headers` order, captured from the live
    # workbook so a created sheet reads identically. Validated against that order
    # in __post_init__ - the two drifting apart is how a column silently loses its
    # width, or worse, how a new fixed column gets missed entirely.
    fixed_col_widths: tuple = (22.55, 9.0, 58.0, 22.55, 15.33, 27.66)

    def fixed_headers(self) -> tuple:
        """The fixed left block, in sheet order.

        The single declaration of that order. Priority and Momentum signal sit
        *inside* the block rather than at its right end because the contact block
        grows rightward and would overwrite anything parked there.
        """
        return (
            self.company_header,
            self.priority_header,
            self.signal_header,
            self.url_header,
            self.industry_header,
            self.status_header,
        )

    def __post_init__(self) -> None:
        if len(self.fixed_col_widths) != len(self.fixed_headers()):
            raise ValueError(
                "fixed_col_widths has {} entries but there are {} fixed columns: {}"
                .format(len(self.fixed_col_widths), len(self.fixed_headers()),
                        ", ".join(self.fixed_headers()))
            )


SHEET_SCHEMA = SheetSchema()


@dataclass
class SheetStructure:
    """Detected column layout of the Company shortlist sheet."""
    header_row: int
    contact_url_row: int  # row holding each contact's own profile URL (header_row + 1)
    data_start_row: int  # first row that is a company (not a URL-ref row)
    company_col: int
    url_col: Optional[int]
    industry_col: Optional[int]
    status_col: Optional[int]
    # Optional like the three above: absent in the older v01 layout, so every call
    # site must guard with `if structure.priority_col:` before writing.
    priority_col: Optional[int] = None
    signal_col: Optional[int] = None
    # contact name → column index
    contact_cols: Dict[str, int] = field(default_factory=dict)

    def company_col_letter(self) -> str:
        return get_column_letter(self.company_col)

    def industry_col_letter(self) -> str:
        return get_column_letter(self.industry_col) if self.industry_col else None

    def status_col_letter(self) -> str:
        return get_column_letter(self.status_col) if self.status_col else None

    def priority_col_letter(self) -> str:
        return get_column_letter(self.priority_col) if self.priority_col else None

    def signal_col_letter(self) -> str:
        return get_column_letter(self.signal_col) if self.signal_col else None

    def contact_col_letters(self) -> Dict[str, str]:
        return {name: get_column_letter(idx) for name, idx in self.contact_cols.items()}

    def last_contact_col_idx(self) -> Optional[int]:
        return max(self.contact_cols.values()) if self.contact_cols else None

    def next_contact_col_idx(self, ws: Worksheet) -> Optional[int]:
        """Return the first column index after the last occupied contact column."""
        if not self.contact_cols:
            # No contacts yet — place right after status col
            start = (self.status_col or self.company_col) + 1
        else:
            start = self.last_contact_col_idx() + 1

        # Scan forward until we find an empty header cell
        for idx in range(start, start + 20):
            cell_val = ws.cell(row=self.header_row, column=idx).value
            if not cell_val:
                return idx
        return None


def detect_sheet_structure(ws: Worksheet, schema: SheetSchema = SHEET_SCHEMA) -> SheetStructure:
    """
    Detect column layout by reading the header row named in ``schema``.

    Fixed columns (Company, Linkedin URL, Industry, Status) are matched by name.
    Contact columns are then detected **positionally**: any non-empty header cell
    to the right of the last fixed column is a contact column, keyed by its bare
    (full-name) header text. Reading (not assuming) the positions keeps the skill
    robust to a user who reordered or renamed-around columns.
    """
    header_row = schema.header_row
    company_col = None
    url_col = None
    industry_col = None
    status_col = None
    priority_col = None
    signal_col = None
    # Header text kept alongside, so a "Priority" column that turns out to sit in the
    # contact block can be handed back to it with its own name intact.
    priority_header_text = None
    signal_header_text = None
    # Non-empty header cells that aren't one of the fixed columns — candidate
    # contact columns, resolved positionally once the anchor is known.
    other_headers: List[Tuple[int, str]] = []  # (col_idx, header_text)

    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col_idx).value
        if not val:
            continue
        val_lower = str(val).lower().strip()
        if val_lower == schema.company_header.lower():
            company_col = col_idx
        elif "linkedin" in val_lower and "url" in val_lower:
            url_col = col_idx
        elif val_lower == schema.industry_header.lower():
            industry_col = col_idx
        elif val_lower == schema.status_header.lower():
            status_col = col_idx
        elif val_lower == schema.priority_header.lower():
            priority_col = col_idx
            priority_header_text = str(val).strip()
        elif val_lower.startswith(schema.signal_header.lower()):
            # Prefix match so a dated variant ("Momentum signal (Aug 2026)") still
            # resolves. The date belongs in the cells, not the column key, but an
            # older workbook should not fail to load over it.
            signal_col = col_idx
            signal_header_text = str(val).strip()
        else:
            other_headers.append((col_idx, str(val).strip()))

    if company_col is None:
        raise ValueError(
            f"Could not find '{schema.company_header}' header in row {header_row}"
        )

    # Anchor for "a contact column sits to the right of the fixed block". Prefer
    # Status; fall back to the rightmost fixed column found if Status is absent.
    anchor_col = max(
        c for c in (status_col, industry_col, url_col, company_col) if c is not None
    )
    # A contact column headed exactly "Priority" (or "Momentum signal...") is a person,
    # not the fixed column of that name: fixed columns live left of the anchor, and
    # anything right of it belongs to the contact block. Matching purely on header text
    # would swallow that column and drop the referrer from contact_cols silently.
    if priority_col is not None and priority_col > anchor_col:
        other_headers.append((priority_col, priority_header_text))
        priority_col = None
    if signal_col is not None and signal_col > anchor_col:
        other_headers.append((signal_col, signal_header_text))
        signal_col = None

    # Keyed by header text. If two contact columns share an identical header the
    # last one wins (dict semantics) — an accepted trade-off of bare full-name
    # headers; the append path guards against creating same-name duplicates.
    contact_cols: Dict[str, int] = {
        header: col_idx
        for col_idx, header in other_headers
        if col_idx > anchor_col
    }

    # data_start_row: the first row after the header holding a company name.
    # The contact-URL row (row 5 in the reference sheet) sits between the header
    # and the data, so scan past any leading URL/blank rows to the first company.
    data_start_row = schema.data_start_row
    for r in range(header_row + 1, header_row + 5):
        val = ws.cell(row=r, column=company_col).value
        if val and not str(val).lower().startswith("http"):
            data_start_row = r
            break

    return SheetStructure(
        header_row=header_row,
        contact_url_row=header_row + 1,  # contacts' own URLs sit just under the header
        data_start_row=data_start_row,
        company_col=company_col,
        url_col=url_col,
        industry_col=industry_col,
        status_col=status_col,
        priority_col=priority_col,
        signal_col=signal_col,
        contact_cols=contact_cols,
    )



def refresh_autofilter(
    ws: Worksheet, structure: SheetStructure, schema: SheetSchema = SHEET_SCHEMA
) -> Optional[str]:
    """Re-span the sheet's autofilter over the full used range.

    Called after any structural change (a company row appended, a contact column
    added) so the filter keeps seeing the whole table. Without this the range stays
    frozen at whatever it was when the filter was first applied, and rows added
    later are silently excluded from filtering and sorting.

    The right edge is the rightmost column the structure actually knows about, not
    ``ws.max_column`` — openpyxl reports trailing empty-but-styled columns, which
    would stretch the filter over blank space. The bottom edge is the last row
    carrying a company name, found by scanning up from the end.

    Returns the range that was set, or ``None`` if there is no data to span.
    """
    known_cols = [
        c for c in (
            structure.company_col, structure.url_col, structure.industry_col,
            structure.status_col, structure.priority_col, structure.signal_col,
        ) if c
    ] + list(structure.contact_cols.values())
    if not known_cols:
        return None

    last_row = None
    for r in range(ws.max_row, structure.data_start_row - 1, -1):
        if ws.cell(row=r, column=structure.company_col).value:
            last_row = r
            break
    if last_row is None:
        return None

    # Both edges resolve from the sheet. Taking the left edge from schema.first_col
    # instead would assume a layout rather than read one, and would start the filter
    # in the wrong place on any sheet whose Company column has moved.
    ref = "{}{}:{}{}".format(
        get_column_letter(min(known_cols)), structure.header_row,
        get_column_letter(max(known_cols)), last_row,
    )
    ws.auto_filter.ref = ref
    return ref

def create_sheet_from_schema(
    contacts: Optional[List[Dict[str, str]]] = None,
    companies: Optional[List[Dict[str, str]]] = None,
    schema: SheetSchema = SHEET_SCHEMA,
) -> Workbook:
    """
    Build a fresh workbook laid out exactly per ``schema``.

    ``contacts`` is an ordered list of ``{"name", "url"}`` dicts — each becomes a
    bare full-name header (row ``header_row``) with the person's own profile
    URL beneath it (row ``contact_url_row``). ``companies`` is an optional list of
    ``{"company", "url", "industry"}`` dicts seeded as data rows from
    ``data_start_row`` with the default status pre-filled.

    The returned workbook is unsaved; the caller persists it. ``init_sheet``
    saves it directly (a brand-new file has no prior data to back up), so this
    path deliberately does not go through :class:`ExcelManager`.
    """
    contacts = contacts or []
    companies = companies or []

    wb = Workbook()
    ws = wb.active
    ws.title = schema.sheet_name

    bold = Font(bold=True)
    header_border = Border(bottom=Side(style="thin"))

    # Title (B2), bold.
    title_cell = ws[schema.title_cell]
    title_cell.value = schema.title_text
    title_cell.font = bold

    # Fixed headers on the header row, starting at the schema's first column.
    # The column each one landed in is remembered so the data below can be written
    # by header rather than by offset.
    fixed_headers = list(schema.fixed_headers())
    fixed_cols: Dict[str, int] = {}
    col = schema.first_col
    for header in fixed_headers:
        cell = ws.cell(row=schema.header_row, column=col)
        cell.value = header
        cell.font = bold
        cell.border = header_border
        fixed_cols[header] = col
        col += 1

    # One contact column per requested introducer: header + profile URL beneath.
    # The header is the contact's bare (full) name — kept in sync with the append
    # path in linkedin_excel_contact_skill._insert_column_with_header.
    for contact in contacts:
        cell = ws.cell(row=schema.header_row, column=col)
        cell.value = contact["name"]
        cell.font = bold
        cell.border = header_border
        url = contact.get("url")
        if url:
            ws.cell(row=schema.contact_url_row, column=col).value = url
        col += 1

    # Seed any starting companies as data rows. Addressed BY HEADER, never by
    # offset from first_col: the fixed block's width is not a constant, and offset
    # arithmetic here would put every value one column left of its label the next
    # time a fixed column is added.
    for offset, company in enumerate(companies):
        row = schema.data_start_row + offset
        seeded = {
            schema.company_header: company.get("company"),
            schema.url_header: company.get("url"),
            schema.industry_header: company.get("industry"),
            schema.status_header: schema.default_status,
            # Rules 4 and 5 travel together: a default rating, and a blank-but-present
            # cell for the justification it still owes.
            schema.priority_header: schema.default_priority,
            schema.signal_header: None,
        }
        for header, value in seeded.items():
            ws.cell(row=row, column=fixed_cols[header]).value = value

    # Column widths: fixed columns from captured values, contact columns uniform.
    for i, width in enumerate(schema.fixed_col_widths):
        ws.column_dimensions[get_column_letter(schema.first_col + i)].width = width
    contact_start = schema.first_col + len(fixed_headers)
    for i in range(len(contacts)):
        ws.column_dimensions[get_column_letter(contact_start + i)].width = schema.contact_col_width

    return wb


def copy_cell_style(source_cell, target_cell) -> None:
    """
    Deep-copy all style attributes from source to target cell.

    Covers: font, border, fill, number_format, alignment, protection.
    """
    if source_cell.has_style:
        target_cell.font = copy.copy(source_cell.font)
        target_cell.border = copy.copy(source_cell.border)
        target_cell.fill = copy.copy(source_cell.fill)
        target_cell.number_format = source_cell.number_format
        target_cell.alignment = copy.copy(source_cell.alignment)
        target_cell.protection = copy.copy(source_cell.protection)


class ExcelManager:
    """Manages Excel file operations with backup and validation."""

    def __init__(self, base_path: str):
        self.base_path = base_path
        self._workbook = None
        self._file_path: Optional[str] = None

    def load_workbook(self, file_path: str):
        """Load, store, and return the workbook. Raises on missing/locked file."""
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        try:
            wb = load_workbook(file_path)
            self._workbook = wb
            self._file_path = file_path
            return wb
        except PermissionError:
            raise PermissionError(
                f"Excel file is locked. Close it in Excel and try again: {file_path}"
            )
        except Exception as e:
            raise RuntimeError(f"Failed to load Excel file: {e}")

    def save_workbook(self, file_path: Optional[str] = None) -> bool:
        """Save the stored workbook. Uses path from load if not provided."""
        path = file_path or self._file_path
        if not path:
            raise RuntimeError("No file path for save — load a workbook first")
        if self._workbook is None:
            raise RuntimeError("No workbook loaded")
        try:
            self._workbook.save(path)
            return True
        except Exception as e:
            raise RuntimeError(f"Failed to save Excel file: {e}")

    def create_backup(self, file_path: str) -> str:
        """
        Create a timestamped crash-safety backup in the ephemeral work dir and
        return its path.

        The backup lives outside the data folder so it never accumulates there;
        it exists purely to restore the original if a save fails, and is deleted
        by :meth:`delete_backup` once the save is confirmed successful.
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Cannot backup non-existent file: {file_path}")
        # Microsecond precision so two runs on the same file within one second
        # don't collide on the same backup name in the shared work dir.
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        name, _ = os.path.splitext(os.path.basename(file_path))
        backup_name = f"{name}_backup_{timestamp}.xlsx"
        backup_path = os.path.join(work_dir(), backup_name)
        shutil.copy2(file_path, backup_path)
        return backup_path

    def restore_backup(self, backup_path: str, target_path: str) -> None:
        """Overwrite target with backup (called on save failure)."""
        if os.path.exists(backup_path):
            shutil.copy2(backup_path, target_path)

    def delete_backup(self, backup_path: str) -> None:
        """Remove the crash-safety backup after a confirmed successful save."""
        if backup_path and os.path.exists(backup_path):
            os.remove(backup_path)

    @property
    def workbook(self):
        return self._workbook
